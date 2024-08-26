from constrainedfilefield.fields import ConstrainedFileField
from django.conf import settings
from django.contrib.auth import get_user_model
from django.db import models
from django.db.models import Max, Q, OuterRef, Subquery, Prefetch, Count, F, ExpressionWrapper, DecimalField, FloatField
from django.db.models.functions import Cast
from django.utils import timezone
from django.utils.functional import cached_property
from django.utils.translation import gettext_lazy as _, get_language
from pyexcel_xlsx import get_data
import json

from clients.models import Organization

User = settings.AUTH_USER_MODEL


def calculate_score(questions):
    """
    An optimized function to calculate the total score of a set of questions.
    :param questions: a query of type Question
    :return: tuple (full_score, total_score)
    """
    weight_subquery = Answer.objects.filter(
        question=OuterRef('pk')
    ).order_by('-weight').values('weight')[:1]

    questions = questions.annotate(
        weight=Subquery(weight_subquery)
    ).prefetch_related(
        Prefetch(
            'answers',
            queryset=Answer.objects.filter(selected_answer=True),
            to_attr='selected_answers'
        )
    )

    total = 0.0
    full_score = 0.0

    for question in questions:
        is_answered = bool(question.selected_answers)
        score = question.selected_answers[0].weight if is_answered else None
        is_scored = score is not None

        if is_scored:
            if question.is_bonus:
                if score:
                    total += score
                    full_score += question.weight
            else:
                total += score
                full_score += question.weight

    return full_score, total


class Section(models.Model):
    # region fields
    title_ar = models.CharField(
        _('Title (AR)'),
        max_length=256,
        blank=True,
    )

    title_en = models.CharField(
        _('Title (EN)'),
        max_length=256,
        blank=False,
    )

    sub_of = models.ForeignKey(
        "Section",
        null=True,
        blank=True,
        on_delete=models.SET_NULL,
        verbose_name=_('Sub of'),
        related_name='sub_sections',
    )

    display_order = models.PositiveSmallIntegerField(
        _('Display Order'),
        null=True,
        blank=True,
    )

    show_flag = models.BooleanField(
        _('Show?'),
        default=True,
        blank=True,
    )

    # endregion fields

    class Meta:
        ordering = ('display_order', 'title_ar', 'title_en')

    @property
    def root_section(self):
        root_section = self
        while root_section.sub_of:
            root_section = root_section.sub_of
        return root_section

    def full_path(self, with_root=True):
        section = self
        full_path = section.title()
        while section.sub_of:
            if not with_root:
                if section.sub_of.sub_of:
                    full_path = section.sub_of.title() + " | " + full_path
            else:
                full_path = section.sub_of.title() + " | " + full_path
            section = section.sub_of

        return full_path

    def questions_of_subsections(self, audit):
        all_sub_sections = Section.objects.filter(Q(pk=self.pk)
                                                  | Q(sub_of=self)
                                                  | Q(sub_of__sub_of=self)
                                                  | Q(sub_of__sub_of__sub_of=self))

        return Question.objects.filter(audit=audit, section__in=all_sub_sections)

    def full_score(self, audit):
        return calculate_score(self.questions_of_subsections(audit=audit))[0]

    def total_score(self, audit):
        return calculate_score(self.questions_of_subsections(audit=audit))[1]

    def weighted_total(self, audit):
        full_score, total_score = calculate_score(self.questions_of_subsections(audit=audit))
        if full_score:
            return total_score / full_score * 100
        else:
            return 0

    weighted_total.short_description = _('Weighted Total (%)')

    def score_in_words(self, audit):
        from django.contrib.humanize.templatetags.humanize import number_format
        full_score, total_score = calculate_score(self.questions_of_subsections(audit=audit))
        if total_score:
            return _('{total_score} out of {full_score} (%{weighted_total})').format(
                total_score=total_score,
                full_score=full_score,
                weighted_total=number_format(total_score/full_score*100, decimal_pos=2),
            )
        else:
            return _('No Score')

    def __str__(self):
        if get_language() == 'ar':
            return self.title_ar or self.title_en or ''
        else:
            return self.title_en or self.title_ar or ''

    def title(self):
        return str(self)


class Audit(models.Model):
    class Types(models.TextChoices):
        RESTAURANTS = 'RESTAURANTS', _('Restaurants')
        HOTELS = 'HOTELS', _('Hotels')
        PILGRIMS_CENTER = 'PILGRIMS_CENTER', _('Pilgrims Center')
        MISC = 'MISC', _('Miscellaneous')

    class Status(models.TextChoices):
        TEMPLATE = 'TEMPLATE', _('Template')
        DRAFT = 'DRAFT', _('Draft')
        SUBMITTED = 'SUBMITTED', _('Submitted')

    # region fields
    title_ar = models.CharField(
        _('Title (AR)'),
        max_length=512,
        blank=True,
        help_text=_('Write a meaningful and descriptive title (or name) for this audit in Arabic.'),
    )

    title_en = models.CharField(
        _('Title (EN)'),
        max_length=512,
        blank=False,
        unique=False,
        help_text=_('Write a meaningful and descriptive title (or name) for this audit in English.'),
    )

    description_ar = models.TextField(
        _('Description (AR)'),
        blank=True,
    )

    description_en = models.TextField(
        _('Description (EN)'),
        blank=True,
    )

    type = models.CharField(
        _('Type'),
        blank=True,
        max_length=128,
        choices=Types.choices,
    )

    created_for = models.ForeignKey(
        "clients.Organization",
        null=True,
        blank=False,
        on_delete=models.SET_NULL,
        verbose_name=_('Created For'),
        related_name='audits',
    )

    inspection_date = models.DateTimeField(
        _('Inspection Date'),
        blank=True,
        null=True,
        help_text=_('Date/time of the inspection visit'),
    )

    derived_from = models.ForeignKey(
        "Audit",
        null=True,
        blank=True,
        on_delete=models.SET_NULL,
        verbose_name=_('Derived From'),
        related_name='driven_audits',
    )

    status = models.CharField(
        _('Status'),
        blank=False,
        max_length=128,
        choices=Status.choices,
        default="DRAFT",
    )
    # endregion fields

    # region audit fields
    created_on = models.DateTimeField(
        _('Created On'),
        auto_now_add=True,
        null=True,
        blank=False,
    )

    created_by = models.ForeignKey(
        get_user_model(),
        null=True,
        blank=True,
        on_delete=models.SET_NULL,
        verbose_name=_('Created By'),
        related_name='audits',
    )

    updated_on = models.DateTimeField(
        _('Updated On'),
        auto_now=True,
        null=True,
        blank=False,
    )

    updated_by = models.ForeignKey(
        get_user_model(),
        null=True,
        blank=True,
        on_delete=models.SET_NULL,
        verbose_name=_('Updated By'),
        related_name='updated_audits',
    )

    # endregion audit

    def __str__(self):
        if get_language() == 'ar':
            return self.title_ar or self.title_en or ''
        else:
            return self.title_en or self.title_ar or ''

    def title(self):
        return str(self)

    def full_score(self):
        return calculate_score(self.questions.all())[0]

    def total_score(self):
        return calculate_score(self.questions.all())[1]

    def weighted_total(self):
        full_score, total = calculate_score(self.questions.all())
        if full_score:
            return total / full_score * 100
        else:
            return 0

    weighted_total.short_description = _('Weighted Total (%)')

    def score_in_words(self):
        from django.contrib.humanize.templatetags.humanize import number_format
        full_score, total_score = calculate_score(self.questions.all())
        if total_score:
            return _('{total_score} out of {full_score} (%{weighted_total})').format(
                total_score=total_score,
                full_score=full_score,
                weighted_total=number_format(total_score / full_score * 100, decimal_pos=1),
            )
        else:
            return _("No Score")

    def get_sections(self):
        root_sections = {}
        all_sections_pks = Question.objects.filter(audit=self).order_by('section').values_list(
            'section__pk',
            flat=True,
        ).distinct()

        all_sections = Section.objects.filter(pk__in=all_sections_pks)

        for section in all_sections:
            section.score = section.score_in_words(audit=self)
            section.full_path = section.full_path(with_root=False)
            section.root_section.score = section.root_section.score_in_words(self)
            if section.root_section not in root_sections.keys():
                root_sections[section.root_section] = [section]
            else:
                root_sections[section.root_section].append(section)

        return root_sections

    def questions_count(self):
        return self.questions.count()

    def get_chart_data(self):
        all_sections_pks = Question.objects.filter(audit=self).order_by('section').values_list(
            'section__pk',
            flat=True,
        ).distinct()

        all_sections = Section.objects.filter(pk__in=all_sections_pks)

        labels, scores = [], []
        for section in all_sections:
            labels.append(str(section)[0:40])
            scores.append(section.weighted_total(audit=self))

        return labels, scores

    def created_by_name(self):
        if self.created_by:
            return self.created_by.get_full_name()

    def inspection_date_hijri(self):
        from hijridate import Gregorian
        from django.utils.safestring import mark_safe
        inspection_date = self.inspection_date.date() if self.inspection_date else self.created_on.date()
        return mark_safe(f"{inspection_date} / {Gregorian.fromdate(inspection_date).to_hijri()}")

    inspection_date_hijri.short_description = _('Inspection Date (Hijri)')


class Question(models.Model):
    # region fields
    audit = models.ForeignKey(
        "Audit",
        null=True,
        blank=False,
        on_delete=models.CASCADE,
        verbose_name=_('Audit'),
        related_name='questions',
    )

    prompt_ar = models.TextField(
        _('Prompt (AR)'),
        blank=True,
    )

    prompt_en = models.TextField(
        _('Prompt (EN)'),
        blank=False,
    )

    help_text_ar = models.TextField(
        _('Help Text (AR)'),
        blank=True,
    )

    help_text_en = models.TextField(
        _('Help Text (EN)'),
        blank=True,
    )

    section = models.ForeignKey(
        "Section",
        null=True,
        blank=False,
        on_delete=models.CASCADE,
        verbose_name=_('Section'),
        related_name='questions',
    )

    display_order = models.PositiveSmallIntegerField(
        _('Display Order'),
        null=True,
        blank=True,
    )

    is_bonus = models.BooleanField(
        _('Is a Bonus Question?'),
        default=False,
    )

    # endregion fields

    class Meta:
        ordering = ('audit', 'section', 'display_order', 'prompt_ar', 'prompt_en', )

    def __str__(self):
        if get_language() == 'ar':
            return self.prompt_ar or self.prompt_en or ''
        else:
            return self.prompt_en or self.prompt_ar or ''

    def prompt(self):
        return str(self)

    prompt.short_description = _('Prompt')

    def help_text(self):
        if get_language() == 'ar':
            return self.help_text_ar
        else:
            return self.help_text_en

    def get_the_answer(self):
        if self.answers.filter(selected_answer=True).exists():
            return self.answers.filter(selected_answer=True).first()

    get_the_answer.short_description = _('Selected Answer')

    def is_answered(self):
        return self.answers.filter(selected_answer=True).exists()

    is_answered.boolean = True
    is_answered.short_description = _('Answered?')

    def weight(self):
        return self.answers.aggregate(Max('weight')).get('weight__max', 0.0)

    def score(self):
        if self.is_answered():
            return self.get_the_answer().weight

    def is_scored(self):
        return self.score() is not None

    is_scored.boolean = True
    is_scored.short_description = _('Scored?')

    def answered_by(self):
        answer = self.get_the_answer()
        if answer:
            return answer.answered_by

    def answered_on(self):
        answer = self.get_the_answer()
        if answer:
            return answer.answered_on

    def change_the_answer(self, the_new_answer, answerer):
        all_other_selected_answers = self.answers.filter(
            selected_answer=True,
        ).exclude(pk=the_new_answer.pk)

        if all_other_selected_answers:
            all_other_selected_answers.update(
                answered_by=None,
                answered_on=None,
                selected_answer=False,
            )

        if not the_new_answer.selected_answer:
            the_new_answer.answered_by = answerer
            the_new_answer.answered_on = timezone.now()
            the_new_answer.selected_answer = True
            the_new_answer.save()

    def evidences(self):
        return self.list_of_evidence.filter(type__in=[
            Evidence.Types.VIDEO,
            Evidence.Types.PICTURE,
            Evidence.Types.MISC,
        ])

    def notes(self):
        return self.list_of_evidence.filter(type__in=[
            Evidence.Types.TEXT_NOTE,
        ])


class Answer(models.Model):
    # region fields
    question = models.ForeignKey(
        "Question",
        null=True,
        blank=False,
        on_delete=models.CASCADE,
        verbose_name=_('Question'),
        related_name='answers',
    )

    prompt_ar = models.CharField(
        _('Prompt (AR)'),
        blank=True,
        max_length=256,
    )

    prompt_en = models.CharField(
        _('Prompt (EN)'),
        blank=False,
        max_length=256,
    )

    weight = models.FloatField(
        _('Weight'),
        null=True,
        blank=True,
        help_text=_('Empty wight indicate N/A answers which are not going to be calculated in the total score'),
    )

    selected_answer = models.BooleanField(
        _('Selected?'),
        default=False,
    )

    display_order = models.PositiveSmallIntegerField(
        _('Display Order'),
        null=True,
        blank=True,
    )

    answered_on = models.DateTimeField(
        _('Answered On'),
        null=True,
        blank=True,
    )

    answered_by = models.ForeignKey(
        User,
        null=True,
        blank=True,
        on_delete=models.SET_NULL,
        verbose_name=_('Answered By'),
        related_name='answers',
    )

    # endregion fields

    class Meta:
        ordering = ('question', 'display_order', '-weight',)

    def __str__(self):
        if get_language() == 'ar':
            return self.prompt_ar or self.prompt_en or ''
        else:
            return self.prompt_en or self.prompt_ar or ''

    def prompt(self):
        return str(self)

    def is_best_answer(self):
        if self.weight:
            return self.weight == self.question.weight()


class Evidence(models.Model):
    class Types(models.TextChoices):
        PICTURE = 'PICTURE', _('Picture'),
        VIDEO = 'VIDEO', _('Video'),
        TEXT_NOTE = 'TEXT_NOTE', _('Note (Text only)')
        MISC = 'misc', _('Miscellaneous'),

    # region fields
    question = models.ForeignKey(
        'Question',
        null=True,
        blank=False,
        on_delete=models.CASCADE,
        related_name='list_of_evidence',
    )

    uploaded_file = ConstrainedFileField(
        verbose_name=_('Attachment'),
        null=True,
        blank=True,
        upload_to='evidence',
        content_types=[
            'application/pdf',
            'image/png',
            'image/bmp',
            'image/jpg',
            'image/jpeg',
            'image/gif',
            # TODO: add video file formats like mp4 and avi
        ],
        max_upload_size=2000000,  # 2.0 mb limit
    )

    notes = models.TextField(
        _('Notes'),
        blank=True,
    )

    type = models.CharField(
        _('Type'),
        blank=False,
        max_length=256,
        choices=Types.choices,
    )

    uploaded_on = models.DateTimeField(
        _('Uploaded On'),
        auto_now_add=True,
        null=True,
    )

    uploaded_by = models.ForeignKey(
        User,
        verbose_name=_('Uploaded By'),
        null=True,
        blank=True,
        on_delete=models.SET_NULL,
    )

    # endregion fields

    def __str__(self):
        return '{} {}'.format(self.type, self.question)


class AuditFile(models.Model):
    audit = models.ForeignKey(
        "Audit",
        null=True,
        blank=False,
        on_delete=models.CASCADE,
        verbose_name=_('Audit'),
        related_name='files',
    )
    filename = models.FileField(max_length=512, upload_to='uploads/')

    def multi_section(self):
        data = get_data(self.filename.file)
        print(json.dumps(data["Checklist"][0][4]))
        for row in data["Checklist"]:
            if len(row) == 5:

                main_section, c_main = Section.objects.get_or_create(title_ar=row[0])
                sub_section, c_sub = Section.objects.get_or_create(title_ar=row[1])
                if c_sub:
                    sub_section.sub_of = main_section
                    sub_section.save()

                # section = Section.objects.filter(title_en=row[2]).first()
                # if not section:
                #     section2 = Section.objects.filter(title_en=row[1]).first()
                #     if not section2:
                #         section1 = Section.objects.filter(title_en=row[0]).first()
                #         if not section1:
                #             section1 = Section.objects.create(title_en=row[0], title_ar=row[0])
                #         if row[0] != row[1]:
                #             section2 = Section.objects.create(title_en=row[1], title_ar=row[1], sub_of=section1)
                #         else:
                #             section2 = section1
                #     if row[2] != row[1]:
                #         section = Section.objects.create(title_en=row[2], title_ar=row[2], sub_of=section2)
                #     else:
                #         section = section2

                q = Question.objects.create(
                    prompt_en=row[3], prompt_ar=row[2], section=sub_section, audit=self.audit)

                Answer.objects.create(
                    question=q,
                    prompt_en="Yes",
                    prompt_ar="Yes",
                    weight=1,
                )

                Answer.objects.create(
                    question=q,
                    prompt_en="No",
                    prompt_ar="No",
                    weight=0,
                )

                Answer.objects.create(
                    question=q,
                    prompt_en="N/A",
                    prompt_ar="N/A",
                    weight=None,
                )

    def multi_section_old(self):
        data = get_data(self.filename.file)
        print(json.dumps(data["Checklist"][0][4]))
        for row in data["Checklist"]:
            if len(row) == 5:

                section = Section.objects.filter(title_en=row[2]).first()
                if not section:
                    section2 = Section.objects.filter(title_en=row[1]).first()
                    if not section2:
                        section1 = Section.objects.filter(title_en=row[0]).first()
                        if not section1:
                            section1 = Section.objects.create(title_en=row[0], title_ar=row[0])
                        if row[0] != row[1]:
                            section2 = Section.objects.create(title_en=row[1], title_ar=row[1], sub_of=section1)
                        else:
                            section2 = section1
                    if row[2] != row[1]:
                        section = Section.objects.create(title_en=row[2], title_ar=row[2], sub_of=section2)
                    else:
                        section = section2

                q = Question.objects.create(prompt_en=row[3], section=section, audit=self.audit)
                answer1 = Answer()
                answer1.question = q
                answer1.prompt_en = "Yes"
                answer1.weight = 2
                answer1.selected_answer = row[4] == 2
                answer1.save()
                answer2 = Answer()
                answer2.question = q
                answer2.prompt_en = "No"
                answer2.weight = 0
                answer2.selected_answer = row[4] == 0
                print(row[4] == 0)
                answer2.save()

    def one_section(self):
        data = get_data(self.filename.file)
        print(json.dumps(data))
        for row in data["Checklist"]:
            if len(row) > 2:
                sections = Section.objects.filter(title_en=row[0])
                if len(sections) == 0:
                    section = Section.objects.create(title_en=row[0])
                else:
                    section = sections[0]

                q = Question.objects.create(prompt_en=row[1], section=section, audit=self.audit)
                answer1 = Answer()
                answer1.question = q
                answer1.prompt_en = "Yes"
                answer1.weight = 2
                answer1.selected_answer = row[2] == 2
                answer1.save()
                answer2 = Answer()
                answer2.question = q
                answer2.prompt_en = "No"
                answer2.weight = 0
                answer2.selected_answer = row[2] == 0
                answer2.save()

    def import_excel(self):
        from openpyxl import Workbook, load_workbook
        from shared.utils import remove_bullet_numbering

        workbook = load_workbook(filename=self.filename.file)
        data = get_data(self.filename.file)
        sheet_data = data[workbook.sheetnames[0]]

        # audit = Audit.objects.create(
        #     title_ar='{} - {}'.format(workbook.sheetnames[0], remove_bullet_numbering(sheet_data[2][0], ":")),
        #     title_en='{} - {}'.format(workbook.sheetnames[0], remove_bullet_numbering(sheet_data[2][0], ":")),
        #     type=Audit.Types.PILGRIMS_CENTER,
        #     status=Audit.Status.DRAFT,
        #     created_for=None,
        # )

        for i in range(7, len(sheet_data)-1):
            row = sheet_data[i]
            if len(row) >= 4:
                main_section, c_main = Section.objects.get_or_create(title_ar=remove_bullet_numbering(row[0]))
                sub_section, c_sub = Section.objects.get_or_create(title_ar=remove_bullet_numbering(row[1]))
                if c_sub:
                    sub_section.sub_of = main_section
                    sub_section.save()

                if workbook.worksheets[0].cell(i+1, 3).fill.fgColor.rgb in ["FFA0D565", "FF92D050"]:
                    is_bonus = True
                else:
                    is_bonus = False

                q = Question.objects.create(
                    prompt_ar=remove_bullet_numbering(row[2]),
                    prompt_en=remove_bullet_numbering(row[2]),
                    section=sub_section,
                    audit=self.audit,
                    is_bonus=is_bonus,
                )

                Answer.objects.create(
                    question=q,
                    prompt_en="Compliant",
                    prompt_ar="ممتثل",
                    weight=1,
                    selected_answer=row[3] == "ممتثل",
                )

                Answer.objects.create(
                    question=q,
                    prompt_en="Non-Compliant",
                    prompt_ar="غير ممتثل",
                    weight=0,
                    selected_answer=row[3] == "غير ممتثل",
                )

                Answer.objects.create(
                    question=q,
                    prompt_en="N/A",
                    prompt_ar="N/A",
                    weight=None,
                    selected_answer=row[3] == "لا ينطبق",
                )
            else:
                break

    def save(self):
        self.import_excel()
        return super(AuditFile, self).save()


class Stats:
    @staticmethod
    def get_answers_occurrences(lang, answer_prompt_en):
        if lang == "ar":
            returned_prompt = 'prompt_ar'
        else:
            returned_prompt = 'prompt_en'

        return Question.objects.values(returned_prompt).annotate(
            answer_count=Cast(Count(
                'answers',
                filter=Q(answers__selected_answer=True, answers__prompt_en=answer_prompt_en),
            ), FloatField()),
            selected_count=Cast(Count(
                'answers',
                filter=Q(answers__selected_answer=True),
            ), FloatField()),
            ratio=ExpressionWrapper(
                F('answer_count') / F('selected_count') * 100.00,
                output_field=FloatField(),
            ),
        ).filter(
            answer_count__gt=1,
        ).order_by('-answer_count', '-ratio')

    @staticmethod
    def get_strengths():
        return Stats.get_answers_occurrences(lang='ar', answer_prompt_en='Compliant')

    @staticmethod
    def get_weaknesses():
        return Stats.get_answers_occurrences(lang='ar', answer_prompt_en='Non-Compliant')

    @staticmethod
    def get_top_unanswered_questions():
        return Stats.get_answers_occurrences(lang='ar', answer_prompt_en='N/A')

    @staticmethod
    def users_count():
        return get_user_model().objects.all().count()

    @staticmethod
    def all_organizations_count():
        return Organization.objects.all().count()

    @staticmethod
    def visited_organizations_count():
        return Organization.objects.annotate(visits_count=Count("audits")).filter(visits_count__gt=0).count()

    @staticmethod
    def inspection_visits_count():
        return Audit.objects.all().count()